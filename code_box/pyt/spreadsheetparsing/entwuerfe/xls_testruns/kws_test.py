#!/home/keuch/gits/keuch/code_box/pyt/spreadsheetparsing/entwuerfe/ve/bin/python3
import os, csv, math, xlrd, re, sys, xlwt, calendar, textwrap, itertools, pandas
from natsort import natsorted, ns
from xlwt import Formula
from xlutils import copy as xlcopy
from colorama import Fore as coly, Style as coln
from pandas import Series, DataFrame, ExcelWriter
import datetime
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

sys.path.append(os.path.abspath('/home/keuch/gits/keuch/code_box/pyt/spreadsheetparsing/entwuerfe/pyplots/lib/'))
from ce_funclib import determine_kernzeit as dtkz, continuity_check
######################################################################################################################

def check_cmdline_params():
    if len(sys.argv) != 3:
        print(sys.argv[0])
        print(textwrap.fill("1. Argument muss eine taegliche Agenten(Terminierungs-)statistik sein, oder ein Verzeichnis mit mehreren davon",80))
        print(textwrap.fill("2. Argument ist die Zieldatei, d.h. eine Exceldatei",80))
        print
        print(textwrap.fill("Beispiel mit jetzigem Setup: ./programm[0] ../test_stats/archiv/CE_alle_Agenten_taeglich_2017-03-09.xls agenten_taeglich.xls ",280))
        exit()
    elif not os.path.isfile(sys.argv[1]):
        if os.path.isdir(sys.argv[1]):
            pmode="dir"
            sourcefile_IB = os.path.abspath(sys.argv[1])
            targetfile = os.path.abspath(sys.argv[2])
            print("source inbound:\t" + sourcefile_IB)
            print("target:\t" + targetfile)
            return sourcefile_IB, targetfile, pmode
        else:
            print(sys.argv[1]+" is not a regular file")
            exit()
    elif not os.path.isfile(sys.argv[2]):
        print(sys.argv[2] + " is not a regular file")
        exit()
    else:
        pmode="file"
        sourcefile_IB = os.path.abspath(sys.argv[1])
        targetfile = os.path.abspath(sys.argv[2])
        print("source inbound:\t" + sourcefile_IB)
        print("target:\t" + targetfile)
        return sourcefile_IB, targetfile, pmode

spinner = itertools.cycle(['-', '\\', '|', '/'])
def spinn():
    sys.stdout.write(next(spinner))  # write the next character, hopefully py3
    sys.stdout.flush()                # flush stdout buffer (actual character display)
    sys.stdout.write('\b')

def parsedate_header(daily_sheet_cell): # turn crap date into nice date
    date_crap = daily_sheet_cell.strip() # comes like this from upstream, date always lives at 1,1
    day, mon, yea = date_crap[0:2], date_crap[3:5], date_crap[6:10]
    date_clea=str(day+"."+mon+"."+yea)
    date_objt = datetime.datetime.strptime(date_clea, "%d.%m.%Y") # this is a python datetime.date object
    return date_objt

def parsedate_full(daily_sheet_cell): # turn crap date into nice date
    date_crap = daily_sheet_cell.strip() # comes like this from upstream, date always lives at 1,1
    day, mon, yea, hou, mnt = date_crap[0:2], date_crap[3:5], date_crap[6:10], date_crap[11:13], date_crap[14:16]
    date_clea=str(day+"."+mon+"."+yea+"."+hou+"."+mnt)
    date_objt = datetime.datetime.strptime(date_clea, "%d.%m.%Y.%H.%M") # this is a python datetime.date object
    return date_objt

def get_cw_ma(agent_cell):
    teile = re.compile(r'(^\D)\s(.*)(\b\d[\d\s]*$)')
    agent_id_set = list()
    try:
        stdort = teile.match(agent_cell).group(1).strip()
        kuerzel = teile.match(agent_cell).group(2).strip()
        number = int(teile.match(agent_cell).group(3).strip())
        agent_id_set.extend((stdort,kuerzel,number))
    except:
        print("sth wrong with " + str(agent_cell))
    return agent_id_set

def get_filelist(folder):
    print ("scanning " + str(folder) + " "),
    agentsfiles = dict()
    for i in (s for s in os.listdir(folder) if s.endswith(".xls")):
        spinn()
        datei = os.path.join(folder,i)
        sheet = xlrd.open_workbook(datei, formatting_info=True).sheet_by_index(0)
        if sheet.nrows == 0:
            continue
        if sheet.cell(0,0) and (sheet.cell(0,0).value == "CE_alles_taeglich" or sheet.cell(0,0).value == "Carexpert_Agent_Gesing"):
            #if sheet.cell(0,0) and sheet.cell(0,0).value == "CE_alles_taeglich":
            sheet_date = parsedate_header(sheet.cell(1,1).value).date() # this will be the dictionary key as it is the unique overall key
            agentsfiles[sheet_date] = datei
    print
    return agentsfiles

def read_entries(datei,doe):
    sheet = xlrd.open_workbook(datei, formatting_info=True).sheet_by_index(0)
    out_dict = dict()
    if sheet.nrows < 3:
        print ("that's a file without entries")
        return

    rows = sheet.nrows
    for i in range(4,sheet.nrows-1):
        stamp = parsedate_full(sheet.cell(i,0).value)
        agent = get_cw_ma(sheet.cell(i,1).value)
        year = stamp.year
        month = stamp.month
        day = stamp.day
        weekday = stamp.strftime('%a')
        hour = stamp.hour
        bzeit=dtkz(stamp)
        week = stamp.isocalendar()[1]
        yearweek = stamp.strftime('%Y-%V') # gibt einen string 'jahr-woche' zb 2017-10
        angeboten = sheet.cell(i,3).value
        bearbeitet = sheet.cell(i,4).value
        verloren = angeboten - bearbeitet
        ht = sheet.cell(i,24).value
        tt = sheet.cell(i,29).value
        acw = ht - tt
        primkey = tuple((stamp,agent[1]))
        doe[primkey] = dict()
        o = doe[primkey]
        o["ag"] = agent[1]
        o["lo"] = agent[0]
        o["dt"] = stamp.date()
        o["yy"] = year
        o["mm"] = month
        o["dd"] = day
        o["hh"] = hour
        o["bz"] = bzeit
        o["ww"] = week
        o["yw"] = yearweek
        o["wd"] = weekday
        o["an"] = angeboten
        o["be"] = bearbeitet
        o["vl"] = verloren
        o["tt"] = tt
        o["ht"] = ht
        o["acw"] = acw
        #o["att"] = att
        #o["aht"] = aht
        #o["aacw"] = aacw
        #print(dir(o))
    return doe


def frame2list(dataframe):
    week_as_list = list()
    for agent,row in dataframe.iterrows():
        values = row.tolist()
        values.insert(0,agent[0])
        values.insert(1,agent[1])
        week_as_list.append(values)
    return week_as_list

def week_from_frame(week_num,frame):
    print()
    print('dataframe ziehen fuer '+str(week_num))
    wk_MON=datetime.datetime.strptime(week_num + '-1', "%Y-%W-%w").date() # erster tag (Montag) der zu bearbeitenden Woche
    wk_SUN=datetime.datetime.strptime(week_num + '-0', "%Y-%W-%w").date() # letzter Tag (Sonntag) der Woche

    if not ((dates_in_dir.min() < wk_MON) and (dates_in_dir.max() >= wk_SUN)):
        print('incomplete...')
        return "incomplete"

    kw = (frame[frame.yw == week_num]).copy()
    kw_filt = kw[['ag','lo','bz','be','ht','tt','acw']].copy()

    if kw_filt.loc[kw_filt['bz'] == 'k'].empty:
        kw_filt['kbe'] = 0
        kw_filt['kht'] = 0
        kw_filt['ktt'] = 0
        kw_filt['kacw'] = 0
    else:
        kw_filt.loc[kw_filt['bz'] == 'k', 'kbe'] = kw_filt['be']
        kw_filt.loc[kw_filt['bz'] == 'k', 'kht'] = kw_filt['ht']
        kw_filt.loc[kw_filt['bz'] == 'k', 'ktt'] = kw_filt['tt']
        kw_filt.loc[kw_filt['bz'] == 'k', 'kacw'] = kw_filt['acw']

    if kw_filt.loc[kw_filt['bz'] == 'n'].empty:
        kw_filt['nbe'] = 0
        kw_filt['nht'] = 0
        kw_filt['ntt'] = 0
        kw_filt['nacw'] = 0
    else:
        kw_filt.loc[kw_filt['bz'] == 'n', 'nbe'] = kw_filt['be']
        kw_filt.loc[kw_filt['bz'] == 'n', 'nht'] = kw_filt['ht']
        kw_filt.loc[kw_filt['bz'] == 'n', 'ntt'] = kw_filt['tt']
        kw_filt.loc[kw_filt['bz'] == 'n', 'nacw'] = kw_filt['acw']

    total=kw_filt.groupby(['lo','ag']).sum()
    total.fillna(0, inplace=True)

    total['o_be'] = total['kbe']+total['nbe']
    total['o_ht'] = ((total['kht']+total['nht'])/total['o_be'])/1440
    total['o_tt'] = ((total['ktt']+total['ntt'])/total['o_be'])/1440
    total['o_acw'] = ((total['kacw']+total['nacw'])/total['o_be'])/1440

    total['kht'] = (total['kht']/total['kbe'])/1440
    total['ktt'] = (total['ktt']/total['kbe'])/1440
    total['kacw'] = (total['kacw']/total['kbe'])/1440

    total['nht'] = (total['nht']/total['nbe'])/1440
    total['ntt'] = (total['ntt']/total['nbe'])/1440
    total['nacw'] = (total['nacw']/total['nbe'])/1440

    total.fillna(0, inplace=True)
    total=total[['o_be','o_ht','o_tt','o_acw','kbe','kht','ktt','kacw','nbe','nht','ntt','nacw']].sort_values('o_be',ascending=False)

    sums=total.sum()
    sums.name = ("N","Summe KW " + str(week_num))
    kwk=kw.loc[kw['bz'] == 'k'].copy()
    kwn=kw.loc[kw['bz'] == 'n'].copy()

    sums.o_ht=(kw['ht'].sum()/1440)/kw['be'].sum()
    sums.o_tt=(kw['tt'].sum()/1440)/kw['be'].sum()
    sums.o_acw=(kw['acw'].sum()/1440)/kw['be'].sum()
    sums.kht=(kwk['ht'].sum()/1440)/kwk['be'].sum()
    sums.ktt=(kwk['tt'].sum()/1440)/kwk['be'].sum()
    sums.kacw=(kwk['acw'].sum()/1440)/kwk['be'].sum()
    sums.nht=(kwn['ht'].sum()/1440)/kwn['be'].sum()
    sums.ntt=(kwn['tt'].sum()/1440)/kwn['be'].sum()
    sums.nacw=(kwn['acw'].sum()/1440)/kwn['be'].sum()
    total = total.append(sums)
    #print (total)
    ### dataframe ist hier komplett ###
    ### rueckgabewert moechte ich aber als liste von listen, damit xlwt schreiben kann ###
    week_as_list = frame2list(total)
    return week_as_list

def check_latest_yweek(sheet):
    maxweek_date={datetime.date(2016,1,1):'2016-1'} # leeres dict mit einem defaultwert, falls das sheet noch leer sein sollte
    for row in range (0,s_row-1):
        s = target_workbook.sheet_by_index(0)
        c = s.cell(row,0)
        if c.value == "KW":
            ywk = s.cell(row,1).value            # die benachbarte Zelle hat einen String im Format '2017-10' (Jahr-Woche)
            Y_W = datetime.datetime.strptime(ywk + '-1', "%Y-%W-%w").date()  # datetime.date aus dem gefundenen string machen
            maxweek_date[Y_W] = ywk   # Dictionary wird befuellt mit dem dt.date als key und dem string als value
    maxdate=maxweek_date[max(maxweek_date.keys())] # aus den dt.dates laesst sich das maximum finden, der dazugehoerige value wird der Rueckgabewert
    return maxdate

def write_weeks(week,vals):
    global s_row

    style_head      = xlwt.easyxf('alignment: horiz centre; font: bold on')
    style_string    = xlwt.easyxf('alignment: horiz centre')
    style_times     = xlwt.easyxf('alignment: horiz centre', num_format_str = "MM:SS")

    w_sheet = target_workbook_w.get_sheet(0)
    w_sheet.write(s_row,0,"KW",style_head)
    w_sheet.write(s_row,1,wk,style_head)
    s_row += 1

    num_datasets=len(vals)

    for dataset in range(0,num_datasets):
        data = vals[dataset]
        data_row = len(data)
        for val in range(0,data_row):
            if val in (0,1):
                if dataset == num_datasets:
                    w_sheet.write(s_row,val,data[val],style_head)
                else:
                    w_sheet.write(s_row,val,data[val],style_string)
            elif val == 2:
                w_sheet.write(s_row,val,data[val],style_string)
            elif val == 6:
                w_sheet.write(s_row,val+1,data[val],style_string)
            elif val == 10:
                w_sheet.write(s_row,val+2,data[val],style_string)
            elif val in (3,4,5):
                w_sheet.write(s_row,val,data[val],style_times)
            elif val in (7,8,9):
                w_sheet.write(s_row,val+1,data[val],style_times)
            elif val in (11,12,13):
                w_sheet.write(s_row,val+2,data[val],style_times)
        s_row += 1
    s_row += 1
    target_workbook_w.save(target)

#######################################################
############## END OF FUNCTION DEFINITIONS ############

source,target,pmode = check_cmdline_params()
dict_o_e = dict()
kwmax=pd.read_excel(target,usecols=[0]).loc['KW'].index.max()
lastday=datetime.datetime.strptime(kwmax+"-0", "%Y-%W-%w")
print(kwmax);quit()

## read everything from directory into a dict and create a dataframe from it
if pmode == "dir":
    filelist = get_filelist(source)   # alle files, die Agentendaten beinhalten
    fehltage = continuity_check(filelist)   # liste der gefundenen files auf luecken pruefen
    print('merging file data into one dictionary')
    for k in sorted(filelist.keys()):
        dict_o_e = read_entries(filelist[k],dict_o_e)  # jedes file aus der Liste wird geparst und ins dict_o_e eingefügt
        spinn()

## transform dictionary and create dataframe from it, rearrange dataframe
column_order = ['dt', 'yy', 'tt', 'bz', 'hh', 'dd', 'acw', 'mm', 'ww', 'yw', 'wd', 'lo', 'ag', 'an', 'be', 'vl', 'ht']
do_frame = DataFrame(dict_o_e)
doe_frame=do_frame.T
doe_frame=doe_frame[column_order]

## get some variables from the frame like: available days, years, and a year-week string
years_in_dir = doe_frame.yy.unique()    # numpy.ndarray of year values
monate_in_dir = doe_frame.mm.unique()   # numpy.ndarray of month numbers
dates_in_dir = doe_frame.dt.unique()    # numpy.ndarray of datetime.date objects
yws_in_dir = sorted(doe_frame.yw.unique())      # numpy.ndarray of week numbers

## open the target excel file and workbook sheets for reading and writing
target_workbook = xlrd.open_workbook(target, formatting_info=True)  # this is the file
target_sheet = target_workbook.sheet_by_index(0)
target_workbook_w = xlcopy.copy(target_workbook)                # a copy is needed to write into
s_row = target_sheet.nrows+1

## im Excelfile lesen, welche Daten dort schon vorhanden sind
max_yweek_intarget = check_latest_yweek(target_sheet) # gibt die hoechste gefundene Woche als '2017-10' string zurueck
print("letzte Kalenderwoche in der Excel-Datei: "+max_yweek_intarget)

## liste der anzufuegenden Wochen generieren
if not max_yweek_intarget in yws_in_dir:
    weeks_to_add = yws_in_dir
else:
    ywk_sheet_pos=yws_in_dir.index(max_yweek_intarget)
    weeks_to_add = yws_in_dir[ywk_sheet_pos+1:]
print ('neuere daten aus diesen Wochen gefunden: '+str(weeks_to_add))

## die beizufuegenden wochen parsen und ins workbook schreiben, dann abspeichern
for wk in weeks_to_add:
    week_as_a_list=week_from_frame(wk,doe_frame) # returns a list of lists, but week_from_frame has a dataframe to plot from
    if week_as_a_list == "incomplete":
        print ("incomplete indeed")
        continue
    else:
        print ("# " + str(wk) + " # wird geschrieben")
        write_weeks(wk,week_as_a_list)

## auf luecken hinweisen
print('folgende Tage waren nicht als excel-datei vorhanden '+str(fehltage))



### TODO add summary line to each week
print()
print('Skript anpassen, damit nicht immer alle Rohdaten neu eingelesen werden, sondern nur die noch nicht vorhandenen...')
#pickelframe=DataFrame(dict_o_e)
#pickelframe.to_pickle(agentenpickel.pkl)
