#!/usr/bin/python
import os, csv, math, xlrd, re, sys, openpyxl, calendar, textwrap, itertools, pandas
from natsort import natsorted, ns
from xlwt import Formula
from xlutils import copy as xlcopy
from colorama import Fore as coly, Style as coln
from pandas import Series, DataFrame, ExcelWriter
import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
pandas.options.mode.chained_assignment = None

def check_cmdline_params():
    if len(sys.argv) != 3:
        print(sys.argv[0])
        print(textwrap.fill("1. Argument muss eine taegliche Agenten(Terminierungs-)statistik sein, oder ein Verzeichnis mit mehreren davon",80))
        print(textwrap.fill("2. Argument ist die Zieldatei, d.h. eine Exceldatei",80))
        print
        print(textwrap.fill("Beispiel mit jetzigem Setup: ./programm[0] ../test_stats/archiv/CE_alle_Agenten_taeglich_2017-03-09.xls agenten_taeglich.xls ",280))
        exit()
    elif not os.path.isfile(sys.argv[1]):
        if os.path.isdir(sys.argv[1]):
            pmode="dir"
            sourcefile_IB = os.path.abspath(sys.argv[1])
            targetfile = os.path.abspath(sys.argv[2])
            print("source inbound:\t" + sourcefile_IB)
            print("target:\t" + targetfile)
            return sourcefile_IB, targetfile, pmode
        else:
            print(sys.argv[1]+" is not a regular file")
            exit()
    elif not os.path.isfile(sys.argv[2]):
        print(sys.argv[2] + " is not a regular file")
        exit()
    else:
        pmode="file"
        sourcefile_IB = os.path.abspath(sys.argv[1])
        targetfile = os.path.abspath(sys.argv[2])
        print("source inbound:\t" + sourcefile_IB)
        print("target:\t" + targetfile)
        return sourcefile_IB, targetfile, pmode

def parsedate_header(daily_sheet_cell): # turn crap date into nice date
    date_crap = daily_sheet_cell.strip() # comes like this from upstream, date always lives at 1,1
    day, mon, yea = date_crap[0:2], date_crap[3:5], date_crap[6:10]
    date_clea=str(day+"."+mon+"."+yea)
    date_objt = datetime.datetime.strptime(date_clea, "%d.%m.%Y") # this is a python datetime.date object
    return date_objt

def parsedate_full(daily_sheet_cell): # turn crap date into nice date
    date_crap = daily_sheet_cell.strip() # comes like this from upstream, date always lives at 1,1
    day, mon, yea, hou, mnt = date_crap[0:2], date_crap[3:5], date_crap[6:10], date_crap[11:13], date_crap[14:16]
    date_clea=str(day+"."+mon+"."+yea+"."+hou+"."+mnt)
    date_objt = datetime.datetime.strptime(date_clea, "%d.%m.%Y.%H.%M") # this is a python datetime.date object
    return date_objt

def get_cw_ma(agent_cell):
    teile = re.compile(r'(^\D)\s(.*)(\b\d[\d\s]*$)')
    agent_id_set = list()
    try:
        stdort = teile.match(agent_cell).group(1).strip()
        kuerzel = teile.match(agent_cell).group(2).strip()
        number = int(teile.match(agent_cell).group(3).strip())
        agent_id_set.extend((stdort,kuerzel,number))
    except:
        print("sth wrong with " + str(agent_cell))
    return agent_id_set

def get_filelist(folder):
    print ("scanning " + str(folder) + " "),
    agentsfiles = dict()
    spinner = itertools.cycle(['-', '\\', '|', '/'])
    for i in (s for s in os.listdir(folder) if s.endswith(".xls")):
        sys.stdout.write(spinner.next())  # write the next character
        sys.stdout.flush()                # flush stdout buffer (actual character display)
        sys.stdout.write('\b') 
        datei = os.path.join(folder,i)
        sheet = xlrd.open_workbook(datei, formatting_info=True).sheet_by_index(0)
        if sheet.nrows == 0:
            continue
        if sheet.cell(0,0) and sheet.cell(0,0).value == "CE_alles_taeglich":
            sheet_date = parsedate_header(sheet.cell(1,1).value).date() # this will be the dictionary key as it is the unique overall key
            agentsfiles[sheet_date] = datei
    print
    return agentsfiles

def read_entries(datei,doe):
    sheet = xlrd.open_workbook(datei, formatting_info=True).sheet_by_index(0)
    out_dict = dict()
    if sheet.nrows < 3:
        print ("that's a file without entries")
        return

    rows = sheet.nrows
    for i in range(4,sheet.nrows-1):
        stamp = parsedate_full(sheet.cell(i,0).value)
        agent = get_cw_ma(sheet.cell(i,1).value)
        year = stamp.year
        month = stamp.month
        day = stamp.day
        weekday = stamp.strftime('%a')
        hour = stamp.hour

        if weekday in ("Sat", "Sun"):
            bzeit = "n"
        else:
            if 11 < hour < 20:
                bzeit = "k"
            else:
                bzeit = "n"
        
        week = stamp.isocalendar()[1]
        angeboten = sheet.cell(i,3).value
        bearbeitet = sheet.cell(i,4).value
        verloren = angeboten - bearbeitet
        ht = sheet.cell(i,24).value
        tt = sheet.cell(i,29).value
        acw = ht - tt
        primkey = tuple((stamp,agent[1]))
        doe[primkey] = dict()
        o = doe[primkey]
        o["ag"] = agent[1]
        o["lo"] = agent[0]
        o["dt"] = stamp.date()
        o["yy"] = year
        o["mm"] = month
        o["dd"] = day
        o["hh"] = hour
        o["bz"] = bzeit
        o["ww"] = week
        o["wd"] = weekday
        o["an"] = angeboten
        o["be"] = bearbeitet
        o["vl"] = verloren
        o["tt"] = tt
        o["ht"] = ht
        o["acw"] = acw
        #o["att"] = att
        #o["aht"] = aht
        #o["aacw"] = aacw
    return doe

def week_start_end(year, week):
    d = datetime.date(year,1,1)
    if(d.weekday()>3):
        d = d+datetime.timedelta(7-d.weekday())
    else:
        d = d - datetime.timedelta(d.weekday())
    dlt = datetime.timedelta(days = (week-1)*7)
    return d + dlt,  d + dlt + datetime.timedelta(days=6)

def chk_wk_complete(year='2017',week='1'):
    sta,end=week_start_end(year, week)
    if dates_in_dir.min() < sta and dates_in_dir.max() > end:
        return "complete"
    else:
        return "incomplete"

def week_from_frame(year,week_num,frame):
    if chk_wk_complete(year,week_num) == "incomplete":
        return "incomplete"
    
    kw = (frame[frame.ww == week_num])
    kw_filt = kw[['ag','bz','be','ht','tt','acw']]

    if kw_filt.loc[kw_filt['bz'] == 'k'].empty:
        kw_filt['kbe'] = 0
        kw_filt['kht'] = 0
        kw_filt['ktt'] = 0
        kw_filt['kacw'] = 0
    else:
        kw_filt.loc[kw_filt['bz'] == 'k', 'kbe'] = kw_filt['be']
        kw_filt.loc[kw_filt['bz'] == 'k', 'kht'] = kw_filt['ht']
        kw_filt.loc[kw_filt['bz'] == 'k', 'ktt'] = kw_filt['tt']
        kw_filt.loc[kw_filt['bz'] == 'k', 'kacw'] = kw_filt['acw']

    if kw_filt.loc[kw_filt['bz'] == 'n'].empty:
        kw_filt['nbe'] = 0
        kw_filt['nht'] = 0
        kw_filt['ntt'] = 0
        kw_filt['nacw'] = 0
    else:
        kw_filt.loc[kw_filt['bz'] == 'n', 'nbe'] = kw_filt['be']
        kw_filt.loc[kw_filt['bz'] == 'n', 'nht'] = kw_filt['ht']
        kw_filt.loc[kw_filt['bz'] == 'n', 'ntt'] = kw_filt['tt']
        kw_filt.loc[kw_filt['bz'] == 'n', 'nacw'] = kw_filt['acw']

    total=kw_filt.groupby('ag').sum()
    total.fillna(0, inplace=True)

    total['o_be'] = total['kbe']+total['nbe']
    total['o_ht'] = ((total['kht']+total['nht'])/total['o_be'])/1440
    total['o_tt'] = ((total['ktt']+total['ntt'])/total['o_be'])/1440
    total['o_acw'] = ((total['kacw']+total['nacw'])/total['o_be'])/1440

    total['kht'] = (total['kht']/total['kbe'])/1440
    total['ktt'] = (total['ktt']/total['kbe'])/1440
    total['kacw'] = (total['kacw']/total['kbe'])/1440

    total['nht'] = (total['nht']/total['nbe'])/1440
    total['ntt'] = (total['ntt']/total['nbe'])/1440
    total['nacw'] = (total['nacw']/total['nbe'])/1440

    total.fillna(0, inplace=True)
    total=total[['o_be','o_ht','o_tt','o_acw','kbe','kht','ktt','kacw','nbe','nht','ntt','nacw']]
    ### dataframe ist hier komplett ###
    
    ### rueckgabewert moechte ich aber als liste von listen, damit xlwt schreiben kann ###
   
    week_as_list = list()
    print week_num
    for agent,row in total.iterrows():
        vallist = row.tolist()
        vallist.insert(0,agent)
        week_as_list.append(vallist)
    return week_as_list


############## END OF FUNCTION DEFINITTIONS ############

source,target,pmode = check_cmdline_params()
srow = xlrd.open_workbook(target).sheet_by_index(0).nrows+1
dict_o_e = dict()

if pmode == "dir":
    filelist=get_filelist(source)
    for k in sorted(filelist.keys()):
        dict_o_e = read_entries(filelist[k],dict_o_e)


column_order = ['dt','yy','dd','mm','ww','wd','lo','ag','an','be','vl','ht','tt','acw','bz','hh']
doe_frame = DataFrame(dict_o_e).T[column_order]
dates_in_dir = doe_frame.dt.unique()
years_in_dir = doe_frame.yy.unique()
kws_in_dir = doe_frame.ww.unique()
monate_in_dir = doe_frame.mm.unique()



for yy in years_in_dir:
    for wk in kws_in_dir:
        week_frame=week_from_frame(yy,wk,doe_frame)
        if week_frame == "incomplete":
            continue
        else:
            for i in week_frame:
                print i
        print


print srow


#total.to_excel(writer,'Colors',startrow=srow,header=False)
#writer.save()
#uniq_agents = kw.ag.unique()
#for agent in uniq_agents:
#    suma = kw[['be','tt']][kw.ag == agent].sum()
